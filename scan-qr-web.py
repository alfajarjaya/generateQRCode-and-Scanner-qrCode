# Scanner qr Code berbasis Web
#  
# Data akan tersimpan di file excel { data.xlsx }
# Data juga tersimpan di database { siswa dengan tabel qrcode } 

from flask import Flask, render_template, Response
import cv2
import numpy as np
import openpyxl
from openpyxl import Workbook
import mysql.connector
import os
import PySimpleGUI as sg

app = Flask(__name__)

mysqldb = mysql.connector.connect(host="localhost", user="root", password="", database="siswa", port=3306)
mycursor = mysqldb.cursor()

def process_qr_data(decoded_info):
    qr_data_format = {
        'Nomor Induk': '',
        'Nama': '',
        'Kelas': '',
        'Jurusan': ''
    }

    lines = decoded_info.split('\n')

    for line in lines:
        if ':' in line:
            key, value = line.split(': ', 1)
            qr_data_format[key.strip()] = value.strip()

    return qr_data_format

def scan_and_save_qr_code(frame, worksheet):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    qr_code_detector = cv2.QRCodeDetector()
    retval, decoded_info, points, straight_qrcode = qr_code_detector.detectAndDecodeMulti(gray)

    if retval:
        for i in range(len(points)):
            cx = int(np.mean(points[i][:, 0]))
            cy = int(np.mean(points[i][:, 1]))

            points_int32 = points[i].astype(np.int32)

            cv2.polylines(frame, [points_int32], isClosed=True, color=(0, 255, 0), thickness=2)

            qr_data = process_qr_data(decoded_info[i])
            
            i = f'Nomor Induk : {qr_data["Nomor Induk"]}, \nNama : {qr_data["Nama"]} ,\nKelas : {qr_data["Kelas"]},\nJurusan : {qr_data["Jurusan"]}'
            
            # Print pesan debug
            sg.popup(f'QRCode data : {i}')

            if any(qr_data.values()):
                if worksheet.max_row == 1:
                    worksheet.append(['Nomor Induk', 'Nama', 'Kelas', 'Jurusan'])

                worksheet.append([
                    qr_data['Nomor Induk'],
                    qr_data['Nama'],
                    qr_data['Kelas'],
                    qr_data['Jurusan']
                ])
                
                sql = "INSERT INTO qrcode (Nomor_Induk, Nama, Kelas, Jurusan) VALUES (%s, %s, %s, %s)"
                values = (qr_data['Nomor Induk'], qr_data['Nama'], qr_data['Kelas'], qr_data['Jurusan'])
                mycursor.execute(sql, values)
                mysqldb.commit()

                workbook.save('d:/produktif bu Tya/App/data.xlsx')

                sg.popup('Data telah tersimpan!')

cap = cv2.VideoCapture(0)

@app.route('/')
def index():
    return render_template('index.html')

def generate_frames():
    cap = cv2.VideoCapture(0)
    while True:
        ret, frame = cap.read()
        scan_and_save_qr_code(frame, worksheet)
        _, buffer = cv2.imencode('.jpg', frame)
        frame = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

    cap.release()

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

if __name__ == '__main__':
    if os.path.exists('d:/produktif bu Tya/App/data.xlsx'):
        workbook = openpyxl.load_workbook('d:/produktif bu Tya/App/data.xlsx')
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active

    app.run(debug=True)
