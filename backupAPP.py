# Backup dari Application Scanner qr Code
#  
# Apabila error bisa anda ambil file ini, namun App ini belum 100% jadi.......

import cv2
import openpyxl
from openpyxl import Workbook
import numpy as np
import os
import PySimpleGUI as sg

def process_qr_data(decoded_info):
    # Format data QR code
    qr_data_format = {
        'Nomor Induk': '',
        'Nama': '',
        'Kelas': '',
        'Jurusan': ''
    }

    # Pisahkan data saat di scan dengan baris baru
    lines = decoded_info.split('\n')

    # Isi data QR code ke dalam format yang telah ditentukan
    for line in lines:
        if ':' in line:
            key, value = line.split(': ', 1)
            qr_data_format[key.strip()] = value.strip()

    return qr_data_format

def scan_and_save_qr_code(frame, worksheet):
    # Konversi frame menjadi skala abu-abu
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Gunakan OpenCV untuk membaca QR code
    qr_code_detector = cv2.QRCodeDetector()
    retval, decoded_info, points, straight_qrcode = qr_code_detector.detectAndDecodeMulti(gray)

    # Cek apakah ada QR code yang terdeteksi
    if retval:
        # Iterasi melalui hasil deteksi
        for i in range(len(points)):
            # Ambil koordinat tengah QR code
            cx = int(np.mean(points[i][:, 0]))
            cy = int(np.mean(points[i][:, 1]))

            # Konversi koordinat ke tipe data int32
            points_int32 = points[i].astype(np.int32)

            # Tampilkan garis di sekitar QR code
            cv2.polylines(frame, [points_int32], isClosed=True, color=(0, 255, 0), thickness=2)

            # Proses data QR code
            qr_data = process_qr_data(decoded_info[i])

            # Print pesan debug
            print('QR Code Data:', qr_data)

            # Cek apakah semua data QR code kosong
            if any(qr_data.values()):
                # Jika worksheet kosong, tambahkan judul
                if worksheet.max_row == 1:
                    worksheet.append(['Nomor Induk', 'Nama', 'Kelas', 'Jurusan'])

                # Simpan data ke file Excel
                worksheet.append([
                    qr_data['Nomor Induk'],
                    qr_data['Nama'],
                    qr_data['Kelas'],
                    qr_data['Jurusan']
                ])
                print('Data QR Code Tersimpan di Excel')

                # Simpan workbook ke file Excel
                workbook.save('d:/produktif bu Tya/App/data.xlsx')

                # Tampilkan alert bahwa data sudah masuk ke Excel
                sg.popup('Data telah tersimpan !')

                # Hentikan program setelah proses selesai
                cap.release()
                cv2.destroyAllWindows()
                exit()

# Mengecek apakah file Excel sudah ada
if os.path.exists('d:/produktif bu Tya/App/data.xlsx'):
    # Jika sudah ada, load workbook yang ada
    workbook = openpyxl.load_workbook('d:/produktif bu Tya/App/data.xlsx')
    worksheet = workbook.active
else:
    # Jika belum ada, buat workbook baru
    workbook = Workbook()
    worksheet = workbook.active

# Inisialisasi kamera
cap = cv2.VideoCapture(0)

while True:
    ret, frame = cap.read()

    # Panggil fungsi untuk memindai dan menyimpan QR code
    scan_and_save_qr_code(frame, worksheet)

    # Tampilkan frame kamera
    cv2.imshow('QR Code Scanner', frame)

    # Tekan tombol 'x' untuk keluar
    if cv2.waitKey(1) & 0xFF == ord('x'):
        break

# Tutup kamera
cap.release()
cv2.destroyAllWindows()